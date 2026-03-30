#!/usr/bin/env python3
"""
processor.py
============
Procesa el archivo ipi_man.xlsx del INDEC y genera dos archivos JSON:
  - ../data/processed/ipi_cuadro1.json  (3 series principales)
  - ../data/processed/ipi_cuadro2.json  (17 series sectoriales)

Uso:
    cd my-ipi-dashboard/scripts
    python processor.py [--input ../data/raw/ipi_man.xlsx]
"""

import json
import argparse
import os
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    import pandas as pd
except ImportError:
    print("ERROR: Instalá las dependencias primero:")
    print("  pip install -r requirements.txt")
    sys.exit(1)


# ──────────────────────────────────────────────────────────────────
# Configuración de columnas (índices 0-based en el archivo Excel)
# ──────────────────────────────────────────────────────────────────

CUADRO1_CONFIG = {
    "hoja": "Cuadro 1",
    "fila_inicio": 8,          # fila 9 en Excel → índice 8
    "columna_fecha": 0,        # columna A (índice 0)
    "series": {
        "nivelGeneral":      3,   # columna D
        "desestacionalizada": 7,  # columna H
        "tendenciaCiclo":    10,  # columna K
    },
}

CUADRO2_CONFIG = {
    "hoja": "Cuadro 2",
    "fila_inicio": 6,          # fila 7 en Excel → índice 6
    "columna_fecha": 0,        # columna A (índice 0)
    "series": {
        "ipiManufacturero":       3,   # D
        "alimentosBebidas":       4,   # E
        "tabaco":                18,   # S
        "textiles":              21,   # V
        "prendasCueroCalzado":   26,   # AA
        "maderaPapel":           30,   # AE
        "petroleo":              34,   # AI
        "quimicos":              40,   # AO
        "cauchoPlastico":        49,   # AX
        "mineralesNoMetalicos":  53,   # BB
        "metalicasBasicas":      60,   # BI
        "productosMetal":        64,   # BM
        "maquinariaEquipo":      68,   # BQ
        "otrosEquipos":          73,   # BV
        "vehiculosAutomotores":  77,   # BZ
        "otroTransporte":        81,   # CD
        "muebles":               84,   # CG
    },
}


# ──────────────────────────────────────────────────────────────────
# Funciones auxiliares
# ──────────────────────────────────────────────────────────────────

def parse_fecha(valor) -> str | None:
    """Convierte un valor de celda fecha a string 'YYYY-MM'."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m")
    if isinstance(valor, str):
        # Intenta parsear formatos comunes
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%Y-%m"):
            try:
                return datetime.strptime(valor.strip(), fmt).strftime("%Y-%m")
            except ValueError:
                continue
    return None


def parse_valor(valor) -> float | None:
    """Convierte un valor de celda numérico a float."""
    if valor is None:
        return None
    try:
        v = float(valor)
        return round(v, 4) if v == v else None  # NaN check
    except (ValueError, TypeError):
        return None


def leer_hoja(wb: openpyxl.Workbook, config: dict) -> dict:
    """
    Lee una hoja del workbook según la configuración dada.
    Retorna un dict con:
      {
        "nombre_serie": [{"fecha": "YYYY-MM", "valor": float|null}, ...],
        ...
      }
    """
    nombre_hoja = config["hoja"]

    # Buscar hoja por nombre (case-insensitive, parcial)
    hoja = None
    for nombre in wb.sheetnames:
        if config["hoja"].lower() in nombre.lower():
            hoja = wb[nombre]
            break

    if hoja is None:
        print(f"  ⚠  Hoja '{nombre_hoja}' no encontrada. Hojas disponibles: {wb.sheetnames}")
        return {}

    print(f"  ✓ Procesando hoja '{hoja.title}'")

    filas = list(hoja.iter_rows(values_only=True))
    fila_inicio = config["fila_inicio"]
    col_fecha = config["columna_fecha"]

    resultado: dict[str, list] = {k: [] for k in config["series"]}

    filas_procesadas = 0
    filas_con_datos = 0

    for fila in filas[fila_inicio:]:
        fecha = parse_fecha(fila[col_fecha]) if col_fecha < len(fila) else None

        if fecha is None:
            # Fin de datos o fila vacía
            break

        filas_procesadas += 1
        tiene_dato = False

        for nombre_serie, col_idx in config["series"].items():
            valor = parse_valor(fila[col_idx]) if col_idx < len(fila) else None
            resultado[nombre_serie].append({"fecha": fecha, "valor": valor})
            if valor is not None:
                tiene_dato = True

        if tiene_dato:
            filas_con_datos += 1

    print(f"    → {filas_procesadas} filas leídas, {filas_con_datos} con datos")

    # Detectar período
    fechas_con_datos = [
        p["fecha"] for serie in resultado.values()
        for p in serie if p["valor"] is not None
    ]
    if fechas_con_datos:
        print(f"    → Período: {min(fechas_con_datos)} a {max(fechas_con_datos)}")

    return resultado


def validar_datos(datos: dict, nombre: str) -> bool:
    """Valida integridad básica de los datos."""
    errores = 0

    for serie, puntos in datos.items():
        if not puntos:
            print(f"  ⚠  [{nombre}] Serie '{serie}' vacía")
            errores += 1
            continue

        nulos = sum(1 for p in puntos if p["valor"] is None)
        pct_nulos = nulos / len(puntos) * 100
        if pct_nulos > 30:
            print(f"  ⚠  [{nombre}] Serie '{serie}': {pct_nulos:.0f}% de valores nulos")

        # Verificar consistencia de fechas (deben ser mensuales)
        fechas = [p["fecha"] for p in puntos]
        duplicados = len(fechas) - len(set(fechas))
        if duplicados > 0:
            print(f"  ⚠  [{nombre}] Serie '{serie}': {duplicados} fechas duplicadas")
            errores += 1

    return errores == 0


# ──────────────────────────────────────────────────────────────────
# Script principal
# ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Procesa ipi_man.xlsx → JSONs")
    parser.add_argument(
        "--input",
        default=str(Path(__file__).parent.parent / "data" / "raw" / "ipi_man.xlsx"),
        help="Ruta al archivo ipi_man.xlsx",
    )
    parser.add_argument(
        "--output-dir",
        default=str(Path(__file__).parent.parent / "data" / "processed"),
        help="Directorio de salida para los JSONs",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)

    print(f"\n{'='*60}")
    print("  IPI Manufacturero · Procesador de datos")
    print(f"{'='*60}")
    print(f"\n  Archivo: {input_path}")

    if not input_path.exists():
        print(f"\n  ERROR: No se encontró el archivo '{input_path}'")
        print("  Copiá ipi_man.xlsx a my-ipi-dashboard/data/raw/")
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)

    # Cargar workbook
    print("\n  Cargando workbook...")
    try:
        # openpyxl no soporta .xls, usamos xlrd a través de pandas si es necesario
        if input_path.suffix.lower() == ".xls":
            # Convertir con pandas + xlrd para .xls legacy
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                # Intentar con xlrd primero
                try:
                    xls = pd.ExcelFile(str(input_path), engine="xlrd")
                    print(f"  ✓ Hojas encontradas: {xls.sheet_names}")
                    # Convertir a openpyxl-compatible via pandas
                    _process_with_pandas(input_path, output_dir)
                    return
                except Exception as e:
                    print(f"  ℹ  xlrd: {e}")
                    print("  Intentando con openpyxl...")

        wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)
        print(f"  ✓ Hojas encontradas: {wb.sheetnames}")

    except Exception as e:
        print(f"\n  ERROR al cargar el archivo: {e}")
        print("  Asegurate de tener instalado xlrd para archivos .xls:")
        print("  pip install xlrd==1.2.0")
        sys.exit(1)

    # ── Procesar Cuadro 1 ──
    print("\n  [Cuadro 1] Series principales...")
    c1_data = leer_hoja(wb, CUADRO1_CONFIG)

    if not c1_data:
        print("  ERROR: No se pudieron leer datos del Cuadro 1")
        sys.exit(1)

    validar_datos(c1_data, "Cuadro 1")

    # Detectar período
    fechas_c1 = sorted(set(
        p["fecha"] for serie in c1_data.values() for p in serie if p["valor"] is not None
    ))

    cuadro1_json = {
        **c1_data,
        "periodoInicio": fechas_c1[0] if fechas_c1 else "2016-01",
        "periodoFin": fechas_c1[-1] if fechas_c1 else "2024-12",
    }

    out_c1 = output_dir / "ipi_cuadro1.json"
    with open(out_c1, "w", encoding="utf-8") as f:
        json.dump(cuadro1_json, f, ensure_ascii=False, indent=2)
    print(f"  ✓ Guardado: {out_c1}")

    # ── Procesar Cuadro 2 ──
    print("\n  [Cuadro 2] Series sectoriales...")
    c2_data = leer_hoja(wb, CUADRO2_CONFIG)

    if not c2_data:
        print("  ERROR: No se pudieron leer datos del Cuadro 2")
        sys.exit(1)

    validar_datos(c2_data, "Cuadro 2")

    out_c2 = output_dir / "ipi_cuadro2.json"
    with open(out_c2, "w", encoding="utf-8") as f:
        json.dump(c2_data, f, ensure_ascii=False, indent=2)
    print(f"  ✓ Guardado: {out_c2}")

    print(f"\n{'='*60}")
    print("  ✓ Proceso completado con éxito")
    print(f"{'='*60}\n")


def _process_with_pandas(input_path: Path, output_dir: Path):
    """
    Alternativa usando pandas + xlrd para archivos .xls legacy.
    """
    import pandas as pd
    import warnings
    warnings.filterwarnings("ignore")

    print("\n  Procesando con pandas+xlrd...")

    def read_series(sheet_name: str, config: dict) -> dict:
        try:
            df = pd.read_excel(
                str(input_path),
                sheet_name=sheet_name,
                header=None,
                engine="xlrd",
            )
        except Exception:
            # Buscar hoja por nombre parcial
            xls = pd.ExcelFile(str(input_path), engine="xlrd")
            matching = [s for s in xls.sheet_names if config["hoja"].lower() in s.lower()]
            if not matching:
                print(f"  ⚠  Hoja '{config['hoja']}' no encontrada")
                return {}
            df = pd.read_excel(str(input_path), sheet_name=matching[0], header=None, engine="xlrd")

        result: dict[str, list] = {k: [] for k in config["series"]}
        fila_inicio = config["fila_inicio"]
        col_fecha = config["columna_fecha"]

        for idx in range(fila_inicio, len(df)):
            row = df.iloc[idx]
            fecha_raw = row.iloc[col_fecha] if col_fecha < len(row) else None

            if pd.isna(fecha_raw) or fecha_raw is None:
                break

            # Parsear fecha
            if hasattr(fecha_raw, "strftime"):
                fecha = fecha_raw.strftime("%Y-%m")
            elif isinstance(fecha_raw, str):
                fecha = parse_fecha(fecha_raw)
            elif isinstance(fecha_raw, (int, float)):
                # Número serial de Excel
                try:
                    from openpyxl.utils.datetime import from_excel
                    fecha = from_excel(int(fecha_raw)).strftime("%Y-%m")
                except Exception:
                    fecha = None
            else:
                fecha = None

            if not fecha:
                break

            for nombre_serie, col_idx in config["series"].items():
                if col_idx < len(row):
                    v = row.iloc[col_idx]
                    valor = None if pd.isna(v) else round(float(v), 4)
                else:
                    valor = None
                result[nombre_serie].append({"fecha": fecha, "valor": valor})

        # Info
        fechas = [p["fecha"] for serie in result.values() for p in serie if p["valor"] is not None]
        if fechas:
            print(f"    → Período: {min(fechas)} a {max(fechas)}")
        return result

    # Cuadro 1
    print("\n  [Cuadro 1]")
    c1 = read_series(CUADRO1_CONFIG["hoja"], CUADRO1_CONFIG)
    fechas_c1 = sorted(set(p["fecha"] for s in c1.values() for p in s if p["valor"] is not None))
    cuadro1_json = {**c1, "periodoInicio": fechas_c1[0] if fechas_c1 else "2016-01", "periodoFin": fechas_c1[-1] if fechas_c1 else "2024-12"}
    out_c1 = output_dir / "ipi_cuadro1.json"
    with open(out_c1, "w", encoding="utf-8") as f:
        json.dump(cuadro1_json, f, ensure_ascii=False, indent=2)
    print(f"  ✓ Guardado: {out_c1}")

    # Cuadro 2
    print("\n  [Cuadro 2]")
    c2 = read_series(CUADRO2_CONFIG["hoja"], CUADRO2_CONFIG)
    out_c2 = output_dir / "ipi_cuadro2.json"
    with open(out_c2, "w", encoding="utf-8") as f:
        json.dump(c2, f, ensure_ascii=False, indent=2)
    print(f"  ✓ Guardado: {out_c2}")

    print(f"\n{'='*60}")
    print("  ✓ Proceso completado con éxito")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
